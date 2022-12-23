from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, numbers
import numpy as np
import matplotlib.pyplot as plt
import pdfkit
from jinja2 import Environment, FileSystemLoader
import statistics_futures


class Interface:
    """
    Класс для хранения ввода пользователя

    Attributes:
        file_name (str): Название файла
        profession (str): Название профессии
    """

    def __init__(self):
        """
        Инициализирует класс Interface
        """
        self.city = None
        inputs = Interface.check_inputs()
        self.file_name = inputs[0]
        self.profession = inputs[1]

    @staticmethod
    def check_inputs():
        """
        Считывает ввод пользователя

        Returns
            (str, str): Кортеж строк с результатами ввода
        """
        # file_name = input('Введите название файла: ')
        # profession = input('Введите название профессии: ')
        file_name = 'vacancies_dif_currencies.csv'
        profession = 'Аналитик'
        return file_name, profession

    def get_city(self):
        # self.city = input('Введите название города: ')
        self.city = 'Москва'


class Report:
    """
    Класс для вывода данных статистики в файл pdf

    Attributes:
        inputs (Interface): Данные с вводом пользователя
        statistics (Statistics): Статистика по вакансиям
    """

    def __init__(self, inputs):
        """
        Инициализирует класс Report

        Args:
            inputs (Interface): Данные с вводом пользователя
            statistics (Statistics): Статистика по вакансиям
        """
        self.inputs = inputs

    def generate_excel(self, dicts):
        """
        Генерирует файл excel со статистикой

        Args:
            dicts (list): Список словарей с данными
        """

        def as_text(value):
            """
            Возвращает значения как строку

            Args
                value: Значения для обработки

            Returns
                str: Соответствующее строчное значение
            """
            if value is None:
                return ""
            return str(value)

        def set_headers(sheet, headers):
            """
            Задает заголовки в таблице

            Args:
                sheet (sheet): Таблица
                headers (list): Заголовки
            """
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i + 1, value=header).font = Font(bold=True)

        def set_length(sheet):
            """
            Задает длину ячеек в таблице

            Args:
                sheet (sheet): Таблица
            """
            for column in sheet.columns:
                length = max(len(as_text(cell.value)) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = length + 2

        def set_borders(sheet):
            """
            Задает толщину границ в таблице

            Args:
                sheet (sheet): Таблица
            """
            for column in sheet.columns:
                for cell in column:
                    if cell.value:
                        cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        wb = Workbook()
        years_sheet = wb.active
        years_sheet.title = 'Статистика по годам'
        cities_sheet = wb.create_sheet('Статистика по городам')
        years_headers = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.inputs.profession}',
                         'Количество вакансий', f'Количество вакансий - {self.inputs.profession}']
        cities_headers = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']

        thin = Side(border_style='thin', color='000000')

        set_headers(years_sheet, years_headers)
        set_headers(cities_sheet, cities_headers)

        for year in dicts[0].keys():
            years_sheet.append([year, dicts[0][year], dicts[2][year], dicts[1][year], dicts[3][year]])

        for city in dicts[4].keys():
            cities_sheet.append([city, dicts[4][city]])

        for row, city in enumerate(dicts[5].keys()):
            cities_sheet.cell(column=4, row=row + 2, value=city)
            cell = cities_sheet.cell(column=5, row=row + 2, value=dicts[5][city])
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

        set_length(years_sheet)
        set_length(cities_sheet)

        set_borders(years_sheet)
        set_borders(cities_sheet)

        wb.save('report.xlsx')

    def generate_image(self, dicts):
        """
        Составляет графики со статистикой и сохраняет их в файл png

        Args:
            dicts (list): Список словарей со статистикой
        """
        fig = plt.figure()
        width = 0.4

        x_nums = np.arange(len(dicts[0].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(221)
        ax.set_title('Уровень зарплат по годам')
        ax.bar(x_list1, dicts[0].values(), width, label='средняя з/п')
        ax.bar(x_list2, dicts[2].values(), width, label=f'з/п {self.inputs.profession.lower()}')
        ax.set_xticks(x_nums, dicts[0].keys(), rotation='vertical')
        ax.tick_params(axis='both', labelsize=8)
        ax.legend(fontsize=8, loc='upper left')
        ax.grid(True, axis='y')

        x_nums = np.arange(len(dicts[1].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2
        ax = fig.add_subplot(222)
        ax.set_title('Уровень вакансий по годам')
        ax.bar(x_list1, dicts[1].values(), width, label='Количество вакансий')
        ax.bar(x_list2, dicts[3].values(), width, label=f'Количество вакансий\n{self.inputs.profession.lower()}')
        ax.set_xticks(x_nums, dicts[1].keys(), rotation='vertical')
        ax.tick_params(labelsize=8)
        ax.legend(fontsize=8, loc='upper left')
        ax.grid(True, axis='y')

        y_nums = np.arange(len(dicts[4].keys()))
        labels = [i.replace(' ', '\n').replace('-', '-\n') for i in dicts[4].keys()]
        ax = fig.add_subplot(223)
        ax.set_title('Уровень зарплат по городам')
        ax.barh(y_nums, dicts[4].values(), align='center')
        ax.set_yticks(y_nums, labels)
        ax.tick_params(labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        x_nums = np.concatenate(([1 - sum(dicts[5].values())], list(dicts[5].values())))
        labels = np.concatenate((['Другие'], list(dicts[5].keys())))
        ax = fig.add_subplot(224)
        ax.set_title('Доля вакансий по городам')
        ax.pie(x_nums, labels=labels, textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')
        plt.show()

    def generate_image_by_city(self, dicts):
        """
        Составляет графики со статистикой и сохраняет их в файл png

        Args:
            dicts (list): Список словарей со статистикой
        """
        fig = plt.figure()
        width = 0.4

        x_nums = np.arange(len(dicts[0].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(221)
        ax.set_title('Уровень зарплат по годам\n и региону')
        ax.bar(x_list1, dicts[0].values(), width, label='средняя з/п')
        ax.bar(x_list2, dicts[2].values(), width, label=f'з/п {self.inputs.profession.lower()}')
        ax.set_xticks(x_nums, dicts[0].keys(), rotation='vertical')
        ax.tick_params(axis='both', labelsize=8)
        ax.legend(fontsize=8, loc='upper left')
        ax.grid(True, axis='y')

        x_nums = np.arange(len(dicts[1].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2
        ax = fig.add_subplot(222)
        ax.set_title('Уровень вакансий по годам\n и региону')
        ax.bar(x_list1, dicts[1].values(), width, label='Количество вакансий')
        ax.bar(x_list2, dicts[3].values(), width, label=f'Количество вакансий\n{self.inputs.profession.lower()}')
        ax.set_xticks(x_nums, dicts[1].keys(), rotation='vertical')
        ax.tick_params(labelsize=8)
        ax.legend(fontsize=8, loc='upper left')
        ax.grid(True, axis='y')

        y_nums = np.arange(len(dicts[4].keys()))
        labels = [i.replace(' ', '\n').replace('-', '-\n') for i in dicts[4].keys()]
        ax = fig.add_subplot(223)
        ax.set_title('Уровень зарплат по городам')
        ax.barh(y_nums, dicts[4].values(), align='center')
        ax.set_yticks(y_nums, labels)
        ax.tick_params(labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        x_nums = np.concatenate(([1 - sum(dicts[5].values())], list(dicts[5].values())))
        labels = np.concatenate((['Другие'], list(dicts[5].keys())))
        ax = fig.add_subplot(224)
        ax.set_title('Доля вакансий по городам')
        ax.pie(x_nums, labels=labels, textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph_pandas.png')
        plt.show()

    def generate_pdf(self, dicts):
        """
        Генерирует файл pdf на основе данных из таблиц и графиков

        Args:
            dicts (list): Словари с данными
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("report.html")
        options = {'enable-local-file-access': None}

        headings = ['Год', 'Средняя<br>зарплата', f'Средняя зарплата -<br>{self.inputs.profession}',
                    'Количество<br>вакансий', f'Количество вакансий -<br>{self.inputs.profession}']
        headings2 = ['Город', 'Уровень зарплат', 'Доля вакансий']
        share_of_vacancies = {k: f'{round(v * 100, 2)}%'.replace('.', ',') for k, v in dicts[5].items()}

        pdf_template = template.render(
            {'profession': self.inputs.profession, 'headings': headings, 'dicts': dicts, 'headings2': headings2,
             'share_of_vacancies': share_of_vacancies})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def generate_pdf_by_city(self, dicts):
        """
        Генерирует файл pdf на основе данных из таблиц и графиков

        Args:
            dicts (list): Словари с данными
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("report_pandas.html")
        options = {'enable-local-file-access': None}

        headings = ['Год', 'Средняя<br>зарплата', f'Средняя зарплата -<br>{self.inputs.profession}',
                    'Количество<br>вакансий', f'Количество вакансий -<br>{self.inputs.profession}']
        headings2 = ['Город', 'Уровень зарплат', 'Доля вакансий']
        share_of_vacancies = {k: f'{round(v * 100, 2)}%'.replace('.', ',') for k, v in dicts[5].items()}

        pdf_template = template.render(
            {'profession': self.inputs.profession, 'headings': headings, 'dicts': dicts, 'headings2': headings2,
             'share_of_vacancies': share_of_vacancies, 'city': self.inputs.city})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report_pandas.pdf', configuration=config, options=options)


def get_pdf():
    """
    Формирует файл pdf со статистикой
    """
    inputs = Interface()
    statistics = statistics_futures.get_statistics(inputs.file_name, inputs.profession)

    report = Report(inputs)
    report.generate_image(statistics)
    report.generate_pdf(statistics)


def get_pdf_by_city():
    inputs = Interface()
    inputs.get_city()
    statistics = statistics_futures.get_statistics_by_city(inputs.file_name, inputs.profession, inputs.city)

    report = Report(inputs)
    report.generate_image_by_city(statistics)
    report.generate_pdf_by_city(statistics)


get_pdf_by_city()
