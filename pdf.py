import csv
import re
from datetime import datetime
from collections import Counter
import sys
from itertools import islice
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, numbers
import numpy as np
import matplotlib.pyplot as plt
import pdfkit
from jinja2 import Environment, FileSystemLoader


class Salary:
    """
    Класс для представления зарплаты

    Attributes:
        salary_from (int): Нижняя граница оклада
        salary_to (int): Верхняя граница оклада
        salary_currency (str): Валюта оклада
    """

    def __init__(self, salary_from, salary_to, salary_currency):
        """
        Инициализирует объект Salary

        Args:
            salary_from (str): Нижняя граница оклада
            salary_to (str): Верхняя граница оклада
            salary_currency (str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def get_average_salary_rub(self):
        """
        Вычисляет среднюю зарплату и переводит ее в рубли

        Returns:
            float: Средняя зарплата в рублях
        """
        return ((float(self.salary_from) + float(self.salary_to)) / 2) * self.currency_to_rub[
            self.salary_currency]


class Vacancy:
    """
    Класс для представления вакансии

    Attributes:
        name (str): Название вакансии
        area_name (str): Название города
        salary (str): Информация о зарплате
        published_at (str): Время и дата публикации вакансии
    """

    def __init__(self, name, area_name, salary,
                 published_at):
        """
        Инициализирует объект Vacancy

        Args:
            name (str): Название вакансии
            area_name (str): Название города
            salary (str): Информация о зарплате
            published_at (str): Время и дата публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

    def get_year(self):
        """
        Возвращает год публикации вакансии

        Returns
            int: Год публикации вакансии
        """
        return int(datetime.strptime(self.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))


class DataSet:
    """
    Класс для представления данных о вакансиях

    Attributes:
        file_name (str): Название обрабатываемого файла
        vacancies_objects (list<Vacancy>): Список вакансий
    """

    def __init__(self, file_name):
        """
        Инициализирует объект DataSet, создает список вакансий по названию файла

        Args:
             file_name (str): Название файла
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.create_vacancies_objects(self.file_name)

    @staticmethod
    def csv_reader(file_name):
        """
        Считывает строки из файла

        Args:
            file_name (str): Название файла

        Returns
            (list, list): Список заголовков и список с вакансиями
        """
        file_csv = open(file_name, encoding='utf_8_sig')
        reader = csv.reader(file_csv)
        list_data = [i for i in reader]

        if len(list_data) == 0:
            print('Пустой файл')
            sys.exit()
        return list_data[0], [i for i in list_data[1:] if '' not in i and len(i) == len(list_data[0])]

    @staticmethod
    def create_vacancies_objects(file_name):
        """
        Создает список объектов Vacancy по названию файла

        Args:
            file_name (str): Название файла

        Returns
            list<Vacancy>: Список вакансий
        """
        headings, vacancies = DataSet.csv_reader(file_name)
        filtered_vacancies = DataSet.csv_filter(vacancies, headings)
        vacancies_objects = [Vacancy(vacancy['name'],
                                     vacancy['area_name'],
                                     Salary(vacancy['salary_from'], vacancy['salary_to'], vacancy['salary_currency']),
                                     vacancy['published_at']
                                     ) for vacancy in filtered_vacancies]
        return vacancies_objects

    @staticmethod
    def csv_filter(reader, list_naming):
        """
        Фильтрует список вакансий

        Args:
            reader (list): Список строк с вакансиями
            list_naming (list): Названия параметров вакансий

        Returns:
            dict: Словарь отфильтрованных вакансий
        """
        return [{list_naming[i]: DataSet.process_vacancy(vacancy[i]) for i, v in enumerate(vacancy)} for vacancy in
                reader]

    @staticmethod
    def process_vacancy(vacancy):
        """
        Очищает строку вакансии

        Args
         vacancy (str): Строка с данными о вакансии

        Returns
            str: Очищенная строка с данными о вакансии
        """
        return ' '.join(('; '.join(re.sub(re.compile('<.*?>'), '', vacancy).split('\n')).split()))


class Statistics:
    """
    Класс для представления статистики по вакансиям

    Attributes:
        dataset (Dataset): Датасет вакансий
        profession (str): Название профессии
        cities (dict): Словарь вакансий по городам
        profession_dataset (list): Список вакансий с подходящей профессией
        number_of_vacancies (dict): Словарь количества вакансий в каждом году
        number_of_vacancies_by_profession (dict): Словарь количества вакансий в каждом году с определенной профессией
    """

    def __init__(self, dataset, profession):
        """
        Инициализирует объект Statistics, подсчитывает атрибуты cities, profession_dataset, number_of_vacancies, number_of_vacancies_by_profession

        Args:
            dataset (Dataset): Датасет вакансий
            profession (str): Название профессии
        """
        self.dataset = dataset
        self.profession = profession
        self.cities = self.get_cities()
        self.profession_dataset = self.get_vacancies_by_profession(profession)
        self.number_of_vacancies = self.get_number_of_vacancies()
        self.number_of_vacancies_by_profession = self.get_number_of_vacancies_by_profession()

    def get_vacancies_by_profession(self, profession):
        """
        Находит вакансии с определенной профессией

        Args
            profession (str): Название профессии

        Returns
            list: Список подходящих вакансий
        """
        return [vacancy for vacancy in self.dataset.vacancies_objects if profession in vacancy.name]

    def get_salary_level(self):
        """
        Подсчитывает среднюю зарплату в каждом году

        Returns
            dict: Словарь, где каждом году соответствует средняя зарплата в этом году
        """
        salary_level = {v: 0 for v in self.number_of_vacancies.keys()}
        for vacancy in self.dataset.vacancies_objects:
            salary_level[vacancy.get_year()] += vacancy.salary.get_average_salary_rub()
        return {year: int(level / self.number_of_vacancies[year]) for year, level in salary_level.items()}

    def get_cities(self):
        """
        Подсчитывает количество вакансий в каждом городе

        Returns
            dict: Словарь, где каждому городу соответсвует количество вакансий в этом городе
        """
        return {city: vacancies for city, vacancies in
                Counter(v.area_name for v in self.dataset.vacancies_objects).items() if
                vacancies / len(self.dataset.vacancies_objects) >= 0.01}

    def get_number_of_vacancies(self):
        """
        Считает количество вакансий в каждом году

        Returns
            dict: Словарь, где каждому году соотвествует количество вакансий в этом году
        """
        return dict(Counter(v.get_year() for v in self.dataset.vacancies_objects))

    def get_salary_level_by_profession(self):
        """
        Подсчитывает среднюю зарплату в каждом году по определенной профессии

        Returns
            dict: Словарь, где каждому году соответствует средняя зарплата в этом году по определенной профессии
        """
        number_of_vacancies = self.number_of_vacancies_by_profession
        salary_level = {v: 0 for v in number_of_vacancies.keys()}
        for vacancy in self.profession_dataset:
            salary_level[vacancy.get_year()] += vacancy.salary.get_average_salary_rub()

        salary_level = {year: int(level / number_of_vacancies[year]) for year, level in salary_level.items() if
                        number_of_vacancies[year] != 0}
        return salary_level if len(salary_level) > 0 else {year: 0 for year in number_of_vacancies.keys()}

    def get_number_of_vacancies_by_profession(self):
        """
        Подсчитывает количество вакансий в каждом году по определенной профессии

        Returns
            dict: Словарь, где каждому году соответствует количество вакансий в этом году по определенной профессии
        """
        vacancies_by_profession = dict(Counter(v.get_year() for v in self.profession_dataset))
        return vacancies_by_profession if len(vacancies_by_profession) > 0 else {year: 0 for year in
                                                                                 self.number_of_vacancies.keys()}

    def get_salary_level_by_city(self):
        """
        Подсчитывает среднюю зарплату в каждом городе

        Return
            dict: Словарь, где каждому городу соответсвует его средняя зарплата
        """
        salary_level = {city: 0 for city in self.cities.keys()}
        for vacancy in self.dataset.vacancies_objects:
            if vacancy.area_name in salary_level.keys():
                salary_level[vacancy.area_name] += vacancy.salary.get_average_salary_rub()
        return Statistics.get_first_n_elems(dict(
            sorted({city: int(salary / self.cities[city]) for city, salary in salary_level.items()}.items(),
                   key=lambda v: v[1],
                   reverse=True)), 10)

    def get_share_of_vacancies_by_city(self):
        """
        Подсчитывает, какую часть от общего количества вакансий составляют вакансии города

        Returns
            dict: Словарь, где каждому городу соотвествует его часть от общего количества вакансий
        """
        return Statistics.get_first_n_elems(
            dict(sorted({k: round(v / len(self.dataset.vacancies_objects), 4) for k, v in
                         self.cities.items()
                         }.items(),
                        key=lambda v: v[1], reverse=True)), 10)

    def dict_to_output(self, d, start):
        """
        Переводит словарь в строку для вывода

        Args:
            d (dict): Словарь для вывода
            start (str): Подпись для вывода

        Returns
            str: Строка для вывода
        """
        return f'{start}: {d}'

    @staticmethod
    def get_first_n_elems(d, n):
        """
        Возвращает первые n элементов словаря

        Args:
            d (dict): Словарь для обработки
            n (int): Количество необходимых элементов

        Return
            dict: Возвращает укороченный словарь
        """
        return dict(islice(d.items(), n))

    def print_result(self):
        """
        Выводит результаты обработки в консоль
        """
        print(self.dict_to_output(self.get_salary_level(), 'Динамика уровня зарплат по годам'))
        print(self.dict_to_output(self.number_of_vacancies, 'Динамика количества вакансий по годам'))
        print(self.dict_to_output(self.get_salary_level_by_profession(),
                                  'Динамика уровня зарплат по годам для выбранной профессии'))
        print(self.dict_to_output(self.number_of_vacancies_by_profession,
                                  'Динамика количества вакансий по годам для выбранной профессии'))
        print(self.dict_to_output(self.get_salary_level_by_city(), 'Уровень зарплат по городам (в порядке убывания)'))
        print(
            self.dict_to_output(self.get_share_of_vacancies_by_city(), 'Доля вакансий по городам (в порядке убывания)'))


class Interface:
    """
    Класс для хранения ввода пользователя

    Attributes:
        file_name (str): Название файла
        profession (str): Название профессии
    """

    def __init__(self):
        """
        Инициализирует класс Interface
        """
        inputs = Interface.check_inputs()
        self.file_name = inputs[0]
        self.profession = inputs[1]

    @staticmethod
    def check_inputs():
        """
        Считывает ввод пользователя

        Returns
            (str, str): Кортеж строк с результатами ввода
        """
        file_name = input('Введите название файла: ')
        profession = input('Введите название профессии: ')
        # file_name = 'vacancies_by_year.csv'
        # profession = 'Аналитик'
        return file_name, profession


class Report:
    """
    Класс для вывода данных статистики в файл pdf

    Attributes:
        inputs (Interface): Данные с вводом пользователя
        statistics (Statistics): Статистика по вакансиям
    """

    def __init__(self, inputs, statistics):
        """
        Инициализирует класс Report

        Args:
            inputs (Interface): Данные с вводом пользователя
            statistics (Statistics): Статистика по вакансиям
        """
        self.inputs = inputs
        self.statistics = statistics

    def get_dict_list(self):
        """
        Составляет список словарей с нужной статистикой

        Returns
            list: Список словарей с нужной статистикой
        """
        return [self.statistics.get_salary_level(), self.statistics.number_of_vacancies,
                self.statistics.get_salary_level_by_profession(), self.statistics.number_of_vacancies_by_profession,
                self.statistics.get_first_n_elems(self.statistics.get_salary_level_by_city(), 10),
                self.statistics.get_first_n_elems(self.statistics.get_share_of_vacancies_by_city(), 10)]

    def generate_excel(self, dicts):
        """
        Генерирует файл excel со статистикой

        Args:
            dicts (list): Список словарей с данными
        """

        def as_text(value):
            """
            Возвращает значения как строку

            Args
                value: Значения для обработки

            Returns
                str: Соответствующее строчное значение
            """
            if value is None:
                return ""
            return str(value)

        def set_headers(sheet, headers):
            """
            Задает заголовки в таблице

            Args:
                sheet (sheet): Таблица
                headers (list): Заголовки
            """
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i + 1, value=header).font = Font(bold=True)

        def set_length(sheet):
            """
            Задает длину ячеек в таблице

            Args:
                sheet (sheet): Таблица
            """
            for column in sheet.columns:
                length = max(len(as_text(cell.value)) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = length + 2

        def set_borders(sheet):
            """
            Задает толщину границ в таблице

            Args:
                sheet (sheet): Таблица
            """
            for column in sheet.columns:
                for cell in column:
                    if cell.value:
                        cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        wb = Workbook()
        years_sheet = wb.active
        years_sheet.title = 'Статистика по годам'
        cities_sheet = wb.create_sheet('Статистика по городам')
        years_headers = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.inputs.profession}',
                         'Количество вакансий', f'Количество вакансий - {self.inputs.profession}']
        cities_headers = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']

        thin = Side(border_style='thin', color='000000')

        set_headers(years_sheet, years_headers)
        set_headers(cities_sheet, cities_headers)

        for year in dicts[0].keys():
            years_sheet.append([year, dicts[0][year], dicts[2][year], dicts[1][year], dicts[3][year]])

        for city in dicts[4].keys():
            cities_sheet.append([city, dicts[4][city]])

        for row, city in enumerate(dicts[5].keys()):
            cities_sheet.cell(column=4, row=row + 2, value=city)
            cell = cities_sheet.cell(column=5, row=row + 2, value=dicts[5][city])
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

        set_length(years_sheet)
        set_length(cities_sheet)

        set_borders(years_sheet)
        set_borders(cities_sheet)

        wb.save('report.xlsx')

    def generate_image(self, dicts):
        """
        Составляет графики со статистикой и сохраняет их в файл png

        Args:
            dicts (list): Список словарей со статистикой
        """
        fig = plt.figure()
        width = 0.4

        x_nums = np.arange(len(dicts[0].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(221)
        ax.set_title('Уровень зарплат по годам')
        ax.bar(x_list1, dicts[0].values(), width, label='средняя з/п')
        ax.bar(x_list2, dicts[2].values(), width, label=f'з/п {self.inputs.profession.lower()}')
        ax.set_xticks(x_nums, dicts[0].keys(), rotation='vertical')
        ax.tick_params(axis='both', labelsize=8)
        ax.legend(fontsize=8, loc='upper left')
        ax.grid(True, axis='y')

        x_nums = np.arange(len(dicts[1].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2
        ax = fig.add_subplot(222)
        ax.set_title('Уровень вакансий по годам')
        ax.bar(x_list1, dicts[1].values(), width, label='Количество вакансий')
        ax.bar(x_list2, dicts[3].values(), width, label=f'Количество вакансий\n{self.inputs.profession.lower()}')
        ax.set_xticks(x_nums, dicts[1].keys(), rotation='vertical')
        ax.tick_params(labelsize=8)
        ax.legend(fontsize=8, loc='upper left')
        ax.grid(True, axis='y')

        y_nums = np.arange(len(dicts[4].keys()))
        labels = [i.replace(' ', '\n').replace('-', '-\n') for i in dicts[4].keys()]
        ax = fig.add_subplot(223)
        ax.set_title('Уровень зарплат по городам')
        ax.barh(y_nums, dicts[4].values(), align='center')
        ax.set_yticks(y_nums, labels)
        ax.tick_params(labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        x_nums = np.concatenate(([1 - sum(dicts[5].values())], list(dicts[5].values())))
        labels = np.concatenate((['Другие'], list(dicts[5].keys())))
        ax = fig.add_subplot(224)
        ax.set_title('Доля вакансий по городам')
        ax.pie(x_nums, labels=labels, textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')
        plt.show()

    def generate_pdf(self, dicts):
        """
        Генерирует файл pdf на основе данных из таблиц и графиков

        Args:
            dicts (list): Словари с данными
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("report.html")
        options = {'enable-local-file-access': None}

        headings = ['Год', 'Средняя<br>зарплата', f'Средняя зарплата -<br>{self.inputs.profession}',
                    'Количество<br>вакансий', f'Количество вакансий -<br>{self.inputs.profession}']
        headings2 = ['Город', 'Уровень зарплат', 'Доля вакансий']
        print(dicts[5])
        share_of_vacancies = {k: f'{round(v * 100, 2)}%'.replace('.', ',') for k, v in dicts[5].items()}

        pdf_template = template.render(
            {'profession': self.inputs.profession, 'headings': headings, 'dicts': dicts, 'headings2': headings2,
             'share_of_vacancies': share_of_vacancies})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)


def get_pdf():
    """
    Формирует файл pdf со статистикой
    """
    inputs = Interface()
    dataset = DataSet(inputs.file_name)
    statistics = Statistics(dataset, inputs.profession)
    statistics.print_result()

    report = Report(inputs, statistics)
    dict_list = report.get_dict_list()
    report.generate_pdf(dict_list)